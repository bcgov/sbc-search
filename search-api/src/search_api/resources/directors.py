from tempfile import NamedTemporaryFile
from flask import Flask, request, jsonify, send_from_directory, abort


from openpyxl import Workbook
from sqlalchemy import desc, func
from functools import reduce
from sqlalchemy.orm.exc import NoResultFound
from flask import Blueprint
from search_api.models import (
    Corporation,
    CorpOpState,
    CorpState,
    CorpParty,
    CorpName,
    Address,
    Office,
    OfficeType,
    OfficesHeld,
    OfficerType,
    Event,
    _get_corpparty_search_results,
    _add_additional_cols_to_search_results,
    _normalize_addr,
)
from search_api.constants import ADDITIONAL_COLS_ADDRESS, ADDITIONAL_COLS_ACTIVE

API = Blueprint('DIRECTORS_API', __name__, url_prefix='/api/v1/directors')


@API.route('/')
def hello():
    return "Welcome to the director search API."


@API.route('/person/search/')
def corpparty_search():
    args = request.args
    results = _get_corpparty_search_results(args)

    # Total number of results
    # This is waaay to expensive on a large db.
    # total_results = results.count()

    # Pagination
    page = int(args.get("page")) if "page" in args else 1
    results = results.paginate(int(page), 20, False)

    corp_parties = []
    for row in results.items:
        result_fields = [
            'corp_party_id', 'first_nme', 'middle_nme', 'last_nme', 'appointment_dt', 'cessation_dt',
            'corp_num', 'party_typ_cd']

        result_dict = {key: getattr(row, key) for key in result_fields}
        result_dict['corp_party_id'] = int(result_dict['corp_party_id'])

        _add_additional_cols_to_search_results(args, row, result_dict)

        corp_parties.append(result_dict)

    return jsonify({'results': corp_parties})

@API.route('/person/search/export/')
def corpparty_search_export():

    # Query string arguments
    args = request.args
    fields = args.getlist('field')
    additional_cols = args.get('additional_cols')

    # Fetching results
    results = _get_corpparty_search_results(args)

    # Exporting to Excel
    wb = Workbook()

    export_dir = "/tmp"
    with NamedTemporaryFile(mode='w+b', dir=export_dir, delete=True) as f:

        sheet = wb.active

        # Sheet headers (first row)
        _ = sheet.cell(column=1, row=1, value="Person Id")
        _ = sheet.cell(column=2, row=1, value="First Name")
        _ = sheet.cell(column=3, row=1, value="Middle Name")
        _ = sheet.cell(column=4, row=1, value="Last Name")
        _ = sheet.cell(column=5, row=1, value="Appointment Date")
        _ = sheet.cell(column=6, row=1, value="Cessation Date")
        _ = sheet.cell(column=7, row=1, value="Corporation Id")

        if _is_addr_search(fields) or additional_cols == ADDITIONAL_COLS_ADDRESS:
            _ = sheet.cell(column=8, row=1, value="Address")
            _ = sheet.cell(column=9, row=1, value="Postal Code")
        elif additional_cols == ADDITIONAL_COLS_ACTIVE:
            _ = sheet.cell(column=8, row=1, value="Status")


        index = 2
        for row in results:

            # CorpParty.corp_party_id
            _ = sheet.cell(column=1, row=index, value=row[1])
            # CorpParty.first_nme
            _ = sheet.cell(column=2, row=index, value=row[2])
            # CorpParty.middle_nme
            _ = sheet.cell(column=3, row=index, value=row[3])
            # CorpParty.last_nme
            _ = sheet.cell(column=4, row=index, value=row[4])
            # CorpParty.appointment_dt
            _ = sheet.cell(column=5, row=index, value=row[5])
            # CorpParty.cessation_dt
            _ = sheet.cell(column=6, row=index, value=row[6])
            # Corporation.corp_num
            _ = sheet.cell(column=7, row=index, value=row[7])
            # Address.addr_line_1, Address.addr_line_2, Address.addr_line_3
            if _is_addr_search(fields) or additional_cols == ADDITIONAL_COLS_ADDRESS:
                # Address.addr_line_1, Address.addr_line_2, Address.addr_line_3
                _ = sheet.cell(column=8, row=index, value=_merge_corpparty_search_addr_fields(row))
                _ = sheet.cell(column=9, row=index, value=row.postal_cd)
            elif additional_cols == ADDITIONAL_COLS_ACTIVE:
                _ = sheet.cell(column=8, row=index, value=_get_state_typ_cd_display_value(row.state_typ_cd))

            index += 1

        current_date = datetime.datetime.strftime(datetime.datetime.now(), "%Y-%m-%d %H:%M:%S")
        filename = "Director Search Results {date}.xlsx".format(date=current_date)
        full_filename_path = "{dir}/{filename}".format(dir=export_dir, filename=filename)
        wb.save(filename=full_filename_path)

        return send_from_directory(export_dir, filename, as_attachment=True)

@API.route('/person/<id>')
def person(id):
    #try:

    result = (CorpParty.query
        .join(Corporation, Corporation.corp_num == CorpParty.corp_num)
        .add_columns(\
            # CorpParty.corp_party_id,
            # CorpParty.first_nme,
            # CorpParty.middle_nme,
            # CorpParty.last_nme,
            # CorpParty.appointment_dt,
            # CorpParty.cessation_dt,
            # CorpParty.corp_num,
            # CorpParty.delivery_addr_id,
            # CorpParty.party_typ_cd
            Corporation.corp_typ_cd,
            Corporation.admin_email,
            # CorpOpState.state_typ_cd,
            # CorpOpState.full_desc,
        ).filter(CorpParty.corp_party_id==int(id))).one()

    person = result[0]
    result_dict = {}
    name = CorpName.query.filter(CorpName.corp_num == person.corp_num).add_columns(CorpName.corp_nme).filter()[0] #TODO: handle multiple names
    offices = Office.query.filter(Office.corp_num == person.corp_num).all()
    delivery_addr = _normalize_addr(person.delivery_addr_id)
    mailing_addr = _normalize_addr(person.mailing_addr_id)

    states = CorpState.query.filter(
        CorpState.corp_num == person.corp_num,
        CorpState.end_event_id == None).all()

    # TODO : list all, or just the one from the correct time.
    corp_delivery_addr = _normalize_addr(offices[0].delivery_addr_id) if offices else ''
    corp_mailing_addr = _normalize_addr(offices[0].mailing_addr_id) if offices else ''

    # TODO: switch to marshmallow.
    result_dict['corp_party_id'] = int(person.corp_party_id)
    result_dict['first_nme'] = person.first_nme
    result_dict['middle_nme'] = person.middle_nme
    result_dict['last_nme'] = person.last_nme
    result_dict['appointment_dt'] = person.appointment_dt
    result_dict['cessation_dt'] = person.cessation_dt
    result_dict['corp_num'] = person.corp_num
    result_dict['corp_nme'] = name.corp_nme
    result_dict['party_typ_cd'] = person.party_typ_cd
    result_dict['corp_party_email'] = person.email_address
    result_dict['delivery_addr'] = delivery_addr
    result_dict['mailing_addr'] = mailing_addr
    result_dict['corp_delivery_addr'] = corp_delivery_addr
    result_dict['corp_mailing_addr'] = corp_mailing_addr
    result_dict['corp_typ_cd'] = result[1]
    result_dict['corp_admin_email'] = result[2]

    result_dict['states'] = [s.as_dict() for s in states]
    # result_dict['full_desc'] = results[0][14]

    return jsonify(result_dict)



@API.route('/person/officesheld/<corppartyid>')
def officesheld(corppartyid):
    results = (OfficerType.query
            .join(OfficesHeld, OfficerType.officer_typ_cd==OfficesHeld.officer_typ_cd)
            .join(CorpParty, OfficesHeld.corp_party_id == CorpParty.corp_party_id)
            #.join(Address, CorpParty.mailing_addr_id == Address.addr_id)
            .join(Event, Event.event_id == CorpParty.start_event_id)
            .add_columns(
                CorpParty.corp_party_id,
                OfficerType.officer_typ_cd,
                OfficerType.short_desc,
                CorpParty.appointment_dt,
                Event.event_timestmp
            )
            .filter(CorpParty.corp_party_id==int(corppartyid))
        )

    offices = []
    for row in results:
        result_dict = {}

        result_dict['corp_party_id'] = int(row[1])
        result_dict['officer_typ_cd'] = row[2]
        result_dict['short_desc'] = row[3]
        result_dict['appointment_dt'] = row[4]
        result_dict['year'] = row[5].year

        offices.append(result_dict)


    person = CorpParty.query.filter(CorpParty.corp_party_id==int(corppartyid)).one()

    # one or both addr may be null, handle each case.
    if person.delivery_addr_id or person.mailing_addr_id:
        if person.delivery_addr_id and person.mailing_addr_id:
            expr = (CorpParty.delivery_addr_id == person.delivery_addr_id) | \
                (CorpParty.mailing_addr_id == person.mailing_addr_id)
        elif person.delivery_addr_id:
            expr = (CorpParty.delivery_addr_id == person.delivery_addr_id)
        elif person.mailing_addr_id:
            expr = (CorpParty.mailing_addr_id == person.mailing_addr_id)

        same_addr = CorpParty.query.add_columns(
            Event.event_timestmp
        ).filter(expr).join(Event, Event.event_id == CorpParty.start_event_id)
    else:
        same_addr = []

    same_name_and_company = CorpParty.query.add_columns(
        Event.event_timestmp
    ).filter(
        CorpParty.first_nme.ilike(person.first_nme),
        CorpParty.last_nme.ilike(person.last_nme),
        CorpParty.corp_num.ilike(person.corp_num),
    ).join(Event, Event.event_id == CorpParty.start_event_id)


    return jsonify({
        'offices': offices,
        'same_addr': [{**s[0].as_dict(), **{'year':int(s[1].year)}} for s in same_addr if s[0].corp_party_id != int(corppartyid)],
        'same_name_and_company': [{**s[0].as_dict(), **{'year':int(s[1].year)}} for s in same_name_and_company if s[0].corp_party_id != int(corppartyid)],
    })


def _get_state_typ_cd_display_value(state_typ_cd):
    if state_typ_cd == "ACT":
        return "ACT"
    else:
        return "HIS"
