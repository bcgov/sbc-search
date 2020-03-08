from flask import Flask, request, jsonify, send_from_directory, abort
from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from sqlalchemy import desc, func
from functools import reduce
from sqlalchemy.orm.exc import NoResultFound
from flask import Blueprint
from search_api.models import (
    Corporation,
    CorpOpState,
    CorpState,
    CorpParty,
    CorpName,
    Address,
    Office,
    OfficeType,
    OfficesHeld,
    OfficerType,
    Event,
    _get_corporation_search_results,
    _normalize_addr,
    _format_office_typ_cd
)

API = Blueprint('BUSINESSES_API', __name__, url_prefix='/api/v1/businesses')

@API.route('/corporation/search/')
def corporation_search():
    args = request.args
    results = _get_corporation_search_results(args)

    # Total number of results
    # This is waaay to expensive.
    # total_results = results.count()

    # Pagination
    page = int(args.get("page")) if "page" in args else 1
    results = results.paginate(int(page), 20, False)

    corporations = []
    for row in results.items:
        result_dict = {}

        result_fields = ['corp_num', 'corp_nme']

        result_dict = {key: getattr(row, key) for key in result_fields}

        corporations.append(result_dict)

    return jsonify({'results': corporations})

@API.route('/corporation/search/export/')
def corporation_search_export():

    # Query string arguments
    args = request.args

    # Fetching results
    results = _get_corporation_search_results(args)

    # Exporting to Excel
    wb = Workbook()

    export_dir = "/tmp"
    with NamedTemporaryFile(mode='w+b', dir=export_dir, delete=True) as f:

        sheet = wb.active

        # Sheet headers (first row)
        _ = sheet.cell(column=1, row=1, value="Corporation Id")
        _ = sheet.cell(column=2, row=1, value="Corp Name")
        # _ = sheet.cell(column=3, row=1, value="Transition Date")
        # _ = sheet.cell(column=4, row=1, value="Address")
        # _ = sheet.cell(column=5, row=1, value="Postal Code")
        # _ = sheet.cell(column=6, row=1, value="City")
        # _ = sheet.cell(column=7, row=1, value="Province")

        index = 2
        for row in results:
            # Corporation.corp_num
            _ = sheet.cell(column=1, row=index, value=row[0])
            # CorpName.corp_nme
            _ = sheet.cell(column=2, row=index, value=row[1])
            # Corporation.transition_dt
            # _ = sheet.cell(column=3, row=index, value=row[3])
            # # Address.addr_line_1
            # _ = sheet.cell(column=4, row=index, value=row[4])
            # # Address.postal_cd
            # _ = sheet.cell(column=5, row=index, value=row[5])
            # # Address.city
            # _ = sheet.cell(column=6, row=index, value=row[6])
            # # Address.province
            # _ = sheet.cell(column=7, row=index, value=row[7])

            index += 1

        current_date = datetime.datetime.strftime(datetime.datetime.now(), "%Y-%m-%d %H:%M:%S")
        filename = "Corporation Search Results {date}.xlsx".format(date=current_date)
        full_filename_path = "{dir}/{filename}".format(dir=export_dir, filename=filename)
        wb.save(filename=full_filename_path)

        return send_from_directory(export_dir, filename, as_attachment=True)

@API.route('/corporation/<id>')
def corporation(id):

    # TODO: move queries to model class.
    result = (
        Corporation.query
        # .join(CorpState, CorpState.corp_num == Corporation.corp_num)
        # .join(CorpOpState, CorpOpState.state_typ_cd == CorpState.state_typ_cd)
        # .join(Office, Office.corp_num == Corporation.corp_num)
        .add_columns(
            Corporation.corp_num,
            Corporation.transition_dt,
            Corporation.admin_email,
            # Office.mailing_addr_id,
            # Office.office_typ_cd,
            # CorpOpState.state_typ_cd,
            # CorpOpState.full_desc
        )
        # .filter(Office.end_event_id == None)
        # .filter(CorpState.end_event_id == None)
        .filter(Corporation.corp_num == id).one())

    corp = result[0]
    offices = Office.query.filter_by(corp_num=id, end_event_id=None)
    names = CorpName.query.filter_by(corp_num=id).order_by(desc(CorpName.end_event_id))

    output = {}
    # TODO: switch to marshmallow.
    output['corp_num'] = corp.corp_num
    output['transition_dt'] = corp.transition_dt
    output['offices'] = []
    for office in offices:
        output['offices'].append({
            'delivery_addr': _normalize_addr(office.delivery_addr_id),  # TODO: get full address.
            'mailing_addr': _normalize_addr(office.mailing_addr_id),
            'office_typ_cd': _format_office_typ_cd(office.office_typ_cd),
            'email_address': office.email_address
        })

    output['admin_email'] = result[3]
    # output['state_typ_cd'] = result[4]
    # output['full_desc'] = result[5]

    output['NAMES'] = []
    for row in names:
        output['NAMES'].append({
            'name': row.corp_nme
        })

    return jsonify(output)
