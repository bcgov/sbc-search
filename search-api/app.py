from openpyxl import Workbook
from tempfile import NamedTemporaryFile

from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from sqlalchemy import desc, func

from models import Corporation, CorpParty, CorpName, Address, app
CORS(app)
from functools import reduce
from models import (
    Corporation, 
    CorpParty, 
    CorpName, 
    Address, 
    OfficesHeld, 
    OfficerType, 
    app,
)


@app.route('/')
def hello():
    return "Welcome to the director search API.s"


def _get_model_by_field(field_name):

    if field_name in ['FIRST_NME','MIDDLE_NME','LAST_NME','APPOINTMENT_DT','CESSATION_DT']: # CorpParty fields
        return eval('CorpParty')
    elif field_name in ['CORP_NUM']: # Corporation fields
        return eval('Corporation')
    elif field_name in ['CORP_NME']: # CorpName fields
        return eval('CorpName')
    elif field_name in ['ADDR_LINE_1','POSTAL_CD','CITY','PROVINCE']: # Address fields
        return eval('Address')
    
    return None


def _get_filter(field, operator, value):
    
    if field == 'ANY_NME':
        return (_get_filter('FIRST_NME', operator, value)
            | _get_filter('MIDDLE_NME', operator, value)
            | _get_filter('LAST_NME', operator, value))

    model = _get_model_by_field(field)
    
    value = value.lower()
    if model:
        Field = func.lower(getattr(model, field))
        if operator == 'contains':
            return Field.contains(value)
        elif operator == 'exact':
            return Field == value
        elif operator == 'endswith':
            return Field.endswith(value)
        elif operator == 'startswith':
            return Field.startswith(value)
        else:
            raise Exception('invalid operator: {}'.format(operator))
    else:
        raise Exception('invalid field: {}'.format(field))


def _get_sort_field(field_name):

    model = _get_model_by_field(field_name)
    if model:
        return getattr(model, field_name)
    else:
        raise Exception('invalid sort field: {}'.format(field_name))


@app.route('/corporation/search/')
def corporation_search():

    args = request.args

    page = 1
    if "page" in args:
        page = int(args.get("page"))

    if "query" not in args:
        return "No search query was received", 400

    query = args["query"]

    # TODO: move queries to model class.
    results = Corporation.query\
        .join(CorpParty, Corporation.CORP_NUM == CorpParty.CORP_NUM)\
        .join(CorpName, Corporation.CORP_NUM == CorpName.CORP_NUM)
    
    results = results.filter(
        (Corporation.CORP_NUM == query) |
        (CorpName.CORP_NME.contains(query)) |
        (CorpParty.FIRST_NME.contains(query)) |
        (CorpParty.LAST_NME.contains(query)))
    
    results = results.paginate(int(page), 20, False)
    
    # TODO: include corpname and corpparties in serialized child using Marshmallow
    return jsonify({
        'results': [row.as_dict() for row in results.items]
    })


@app.route('/person/search/')
def corpparty_search():
    """
    Querystring parameters as follows:

    You may provide query=<string> for a simple search, OR any number of querystring triples such as

    field=ANY_NME|FIRST_NME|LAST_NME|<any column name>
    &operator=exact|contains|startswith|endswith
    &value=<string>
    &sort_type=asc|desc
    &sort_value=ANY_NME|FIRST_NME|LAST_NME|<any column name>

    For example, to get everyone who has any name that starts with 'Sky', or last name must be exactly 'Little', do:
    curl "http://localhost/person/search/?field=ANY_NME&operator=startswith&value=Sky&field=LAST_NME&operator=exact&value=Little&mode=ALL"
    """

    args = request.args

    page = int(args.get("page")) if "page" in args else 1

    query = args.get("query")

    fields = args.getlist('field')
    operators = args.getlist('operator')
    values = args.getlist('value')
    mode = args.get('mode')
    sort_type = args.get('sort_type')
    sort_value = args.get('sort_value')

    if query and len(fields) > 0:
        raise Exception("use simple query or advanced. don't mix")
    
    # Only triples of clauses are allowed. So, the same number of fields, ops and values.
    if len(fields) != len(operators) or len(operators) != len(values):
        raise Exception("mismatched query param lengths: fields:{} operators:{} values:{}".format(
            len(fields),
            len(operators),
            len(values)))

    # Zip the lists, so ('LAST_NME', 'FIRST_NME') , ('contains', 'exact'), ('Sky', 'Apple') => (('LAST_NME', 'contains', 'Sky'), ('FIRST_NME', 'exact', 'Apple'))
    clauses = list(zip(fields, operators, values))

    # TODO: move queries to model class.
    results = CorpParty.query\
            .filter(CorpParty.END_EVENT_ID == None)\
            .filter(CorpParty.PARTY_TYP_CD.in_(['FIO', 'DIR','OFF']))\
            .filter(CorpName.END_EVENT_ID == None)\
            .join(Corporation, Corporation.CORP_NUM == CorpParty.CORP_NUM)\
            .join(CorpName, Corporation.CORP_NUM == CorpName.CORP_NUM)\
            .join(Address, CorpParty.MAILING_ADDR_ID == Address.ADDR_ID)\
            .add_columns(\
                CorpParty.CORP_PARTY_ID, 
                CorpParty.FIRST_NME, 
                CorpParty.MIDDLE_NME,
                CorpParty.LAST_NME,
                CorpParty.APPOINTMENT_DT,
                CorpParty.CESSATION_DT,
                Corporation.CORP_NUM,
                CorpName.CORP_NME,
                Address.ADDR_LINE_1,
                Address.POSTAL_CD,
                Address.CITY,
                Address.PROVINCE,
            )
    
    # Simple mode - return reasonable results for a single search string:
    if query:
        results = results.filter((Corporation.CORP_NUM == query) | (CorpParty.FIRST_NME.contains(query)) | (CorpParty.LAST_NME.contains(query)))
    # Advanced mode - return precise results for a set of clauses.
    elif clauses:

        # Determine if we will combine clauses with OR or AND. mode=ALL means we use AND. Default mode is OR
        if mode == 'ALL':
            def fn(accumulator, s):
                return accumulator & _get_filter(*s)
        else:
            def fn(accumulator, s):
                return accumulator | _get_filter(*s)
        
        # We use reduce here to join all the items in clauses with the & operator or the | operator.
        # Similar to if we did "|".join(clause), but calling the boolean operator instead.
        filter_grp = reduce(
            fn,
            clauses[1:],
            _get_filter(*clauses[0])
            )
        results = results.filter(filter_grp)

    # Sorting
    if sort_type is None:
        results = results.order_by(CorpParty.LAST_NME)
    else:
        field = _get_sort_field(sort_value)

        if sort_type == 'desc':
            results = results.order_by(desc(field))
        else:
            results = results.order_by(field)
    
    total_results = results.count()

    # Pagination
    results = results.paginate(int(page), 20, False)

    corp_parties = []
    for row in results.items:
        result_dict = {}

        # TODO: switch to marshmallow.
        result_dict['CORP_PARTY_ID'] = row[1]
        result_dict['FIRST_NME'] = row[2]
        result_dict['MIDDLE_NME'] = row[3]
        result_dict['LAST_NME'] = row[4]
        result_dict['APPOINTMENT_DT'] = row[5]
        result_dict['CESSATION_DT'] = row[6]
        result_dict['CORP_NUM'] = row[7]
        result_dict['CORP_NME'] = row[8]
        result_dict['ADDR_LINE_1'] = row[9]
        result_dict['POSTAL_CD'] = row[10]
        result_dict['CITY'] = row[11]
        result_dict['PROVINCE'] = row[12]

        corp_parties.append(result_dict)
    
    return jsonify({'results': corp_parties, 'total': total_results })


@app.route('/person/search/export/')
def corpparty_search_export():

    args = request.args

    page = int(args.get("page")) if "page" in args else 1

    query = args.get("query")

    fields = args.getlist('field')
    operators = args.getlist('operator')
    values = args.getlist('value')
    mode = args.get('mode')
    sort_type = args.get('sort_type')
    sort_value = args.get('sort_value')

    if query and len(fields) > 0:
        raise Exception("use simple query or advanced. don't mix")
    
    # Only triples of clauses are allowed. So, the same number of fields, ops and values.
    if len(fields) != len(operators) or len(operators) != len(values):
        raise Exception("mismatched query param lengths: fields:{} operators:{} values:{}".format(
            len(fields),
            len(operators),
            len(values)))

    # Zip the lists, so ('LAST_NME', 'FIRST_NME') , ('contains', 'exact'), ('Sky', 'Apple') => (('LAST_NME', 'contains', 'Sky'), ('FIRST_NME', 'exact', 'Apple'))
    clauses = list(zip(fields, operators, values))

    # TODO: move queries to model class.
    results = CorpParty.query\
            .filter(CorpParty.END_EVENT_ID == None)\
            .filter(CorpParty.PARTY_TYP_CD.in_(['FIO', 'DIR','OFF']))\
            .filter(CorpName.END_EVENT_ID == None)\
            .join(Corporation, Corporation.CORP_NUM == CorpParty.CORP_NUM)\
            .join(CorpName, Corporation.CORP_NUM == CorpName.CORP_NUM)\
            .join(Address, CorpParty.MAILING_ADDR_ID == Address.ADDR_ID)\
            .add_columns(\
                CorpParty.CORP_PARTY_ID, 
                CorpParty.FIRST_NME, 
                CorpParty.MIDDLE_NME,
                CorpParty.LAST_NME,
                CorpParty.APPOINTMENT_DT,
                CorpParty.CESSATION_DT,
                Corporation.CORP_NUM,
                CorpName.CORP_NME,
                Address.ADDR_LINE_1,
                Address.POSTAL_CD,
                Address.CITY,
                Address.PROVINCE,
            )
    
    # Simple mode - return reasonable results for a single search string:
    if query:
        results = results.filter((Corporation.CORP_NUM == query) | (CorpParty.FIRST_NME.contains(query)) | (CorpParty.LAST_NME.contains(query)))
    # Advanced mode - return precise results for a set of clauses.
    elif clauses:

        # Determine if we will combine clauses with OR or AND. mode=ALL means we use AND. Default mode is OR
        if mode == 'ALL':
            def fn(accumulator, s):
                return accumulator & _get_filter(*s)
        else:
            def fn(accumulator, s):
                return accumulator | _get_filter(*s)
        
        # We use reduce here to join all the items in clauses with the & operator or the | operator.
        # Similar to if we did "|".join(clause), but calling the boolean operator instead.
        filter_grp = reduce(
            fn,
            clauses[1:],
            _get_filter(*clauses[0])
            )
        results = results.filter(filter_grp)

    # Sorting
    if sort_type is None:
        results = results.order_by(CorpParty.LAST_NME)
    else:
        field = _get_sort_field(sort_value)

        if sort_type == 'desc':
            results = results.order_by(desc(field))
        else:
            results = results.order_by(field)

    # Exporting to Excel
    wb = Workbook()

    with NamedTemporaryFile(mode='w+b', dir='tmp', delete=True) as tmp:

        sheet = wb.active

        # Sheet headers (first row)
        _ = sheet.cell(column=1, row=1, value="Person Id")
        _ = sheet.cell(column=2, row=1, value="First Name")
        _ = sheet.cell(column=3, row=1, value="Middle Name")
        _ = sheet.cell(column=4, row=1, value="Last Name")
        _ = sheet.cell(column=5, row=1, value="Appointment Date")
        _ = sheet.cell(column=6, row=1, value="Cessation Date")
        _ = sheet.cell(column=7, row=1, value="Corporation Id")
        _ = sheet.cell(column=8, row=1, value="Corp Name")
        _ = sheet.cell(column=9, row=1, value="Address")
        _ = sheet.cell(column=10, row=1, value="Postal Code")
        _ = sheet.cell(column=11, row=1, value="City")
        _ = sheet.cell(column=12, row=1, value="Province")
        
        index = 2
        for row in results:
            
            # CorpParty.CORP_PARTY_ID
            _ = sheet.cell(column=1, row=index, value=row[1])
            # CorpParty.FIRST_NME
            _ = sheet.cell(column=2, row=index, value=row[2])
            # CorpParty.MIDDLE_NME
            _ = sheet.cell(column=3, row=index, value=row[3])
            # CorpParty.LAST_NME
            _ = sheet.cell(column=4, row=index, value=row[4])
            # CorpParty.APPOINTMENT_DT
            _ = sheet.cell(column=5, row=index, value=row[5])
            # CorpParty.CESSATION_DT
            _ = sheet.cell(column=6, row=index, value=row[6])
            # Corporation.CORP_NUM
            _ = sheet.cell(column=7, row=index, value=row[7])
            # CorpName.CORP_NME
            _ = sheet.cell(column=8, row=index, value=row[8])
            # Address.ADDR_LINE_1
            _ = sheet.cell(column=9, row=index, value=row[9])
            # Address.POSTAL_CD
            _ = sheet.cell(column=10, row=index, value=row[10])
            # Address.CITY
            _ = sheet.cell(column=11, row=index, value=row[11])
            # Address.PROVINCE
            _ = sheet.cell(column=12, row=index, value=row[12])
            
            index += 1

        filename = "{}.{}".format(tmp.name,"xlsx")
        wb.save(filename=filename)

        # file name without the path
        simple_name = filename.split('/')[len(filename.split('/'))-1]        

        return send_from_directory('tmp',simple_name,as_attachment=True)


@app.route('/person/<id>')
def person(id):
    results = CorpParty.query\
            .join(Corporation, Corporation.CORP_NUM == CorpParty.CORP_NUM)\
            .join(CorpName, Corporation.CORP_NUM == CorpName.CORP_NUM)\
            .join(Address, CorpParty.MAILING_ADDR_ID == Address.ADDR_ID)\
            .add_columns(\
                CorpParty.CORP_PARTY_ID, 
                CorpParty.FIRST_NME, 
                CorpParty.MIDDLE_NME,
                CorpParty.LAST_NME,
                CorpParty.APPOINTMENT_DT,
                CorpParty.CESSATION_DT,
                Corporation.CORP_NUM,
                CorpName.CORP_NME,
                Address.ADDR_LINE_1,
                Address.POSTAL_CD,
                Address.CITY,
                Address.PROVINCE,
            ).filter(CorpParty.CORP_PARTY_ID==int(id))
    
    if results.count() > 0:
        result_dict = {}

        # TODO: switch to marshmallow.
        result_dict['CORP_PARTY_ID'] = results[0][1]
        result_dict['FIRST_NME'] = results[0][2]
        result_dict['MIDDLE_NME'] = results[0][3]
        result_dict['LAST_NME'] = results[0][4]
        result_dict['APPOINTMENT_DT'] = results[0][5]
        result_dict['CESSATION_DT'] = results[0][6]
        result_dict['CORP_NUM'] = results[0][7]
        result_dict['CORP_NME'] = results[0][8]
        result_dict['ADDR_LINE_1'] = results[0][9]
        result_dict['POSTAL_CD'] = results[0][10]
        result_dict['CITY'] = results[0][11]
        result_dict['PROVINCE'] = results[0][12]

        return jsonify(result_dict)

    return {}


@app.route('/person/officesheld/<corppartyid>')
def officesheld(corppartyid):
    results = OfficerType.query\
            .join(OfficesHeld, OfficerType.OFFICER_TYP_CD==OfficesHeld.OFFICER_TYP_CD)\
            .join(CorpParty, OfficesHeld.CORP_PARTY_ID == CorpParty.CORP_PARTY_ID)\
            .join(Address, CorpParty.MAILING_ADDR_ID == Address.ADDR_ID)\
            .add_columns(\
                CorpParty.CORP_PARTY_ID, 
                OfficerType.OFFICER_TYP_CD,
                OfficerType.SHORT_DESC,
                CorpParty.APPOINTMENT_DT,
                Address.ADDR_LINE_1,
            )\
            .filter(CorpParty.CORP_PARTY_ID==int(corppartyid))
    
    offices = []
    for row in results:
        result_dict = {}

        result_dict['CORP_PARTY_ID'] = row[1]
        result_dict['OFFICER_TYP_CD'] = row[2]
        result_dict['SHORT_DESC'] = row[3]
        result_dict['APPOINTMENT_DT'] = row[4]
        result_dict['ADDR_LINE_1'] = row[5]

        offices.append(result_dict)
    
    return jsonify({'results': offices})


@app.route('/corporation/<id>')
def corporation(id):
    results = Corporation.query.filter_by(CORP_NUM=id)
    if results.count() > 0:
        return jsonify(results[0].as_dict())
    return {}


if __name__ == '__main__':
    app.run(host='0.0.0.0')
