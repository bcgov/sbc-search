# Copyright © 2020 Province of British Columbia
#
# Licensed under the Apache License, Version 2.0 (the 'License');
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an 'AS IS' BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""API endpoints for searching for and retrieving information about Corporations."""

import datetime
from http import HTTPStatus
from tempfile import NamedTemporaryFile

from flask import Blueprint, request, jsonify, send_from_directory, current_app
from openpyxl import Workbook
from sqlalchemy.sql import literal_column

from search_api.auth import jwt, authorized
from search_api.models.address import Address
from search_api.models.corporation import Corporation
from search_api.models.corp_name import CorpName
from search_api.models.office import Office
from search_api.utils.model_utils import _merge_addr_fields, _format_office_typ_cd
from search_api.utils.utils import convert_to_snake_case


API = Blueprint("BUSINESSES_API", __name__, url_prefix="/api/v1/businesses")


@API.route("/")
@jwt.requires_auth
def corporation_search():
    """Search for Corporations by keyword or corpNum.

    This function takes the following query arguments:
    - query={search keyword}
    - page={page number}
    """
    account_id = request.headers.get("X-Account-Id", None)

    if not authorized(jwt, account_id):
        return (
            jsonify({'message': 'User is not authorized to access Director Search'}),
            HTTPStatus.UNAUTHORIZED,
        )

    # args <- ImmutableMultiDict([('query', 'countable'), ('page', '1'), ('sort_type', 'dsc'), ('sort_value', 'corpNme')])

    args = request.args
    if not args.get("query"):
        return "No search query was received", 400

    # Due to performance issues, exclude address.
    results = Corporation.search_corporations(args, include_addr=False)

    # Pagination
    per_page = 50
    page = int(args.get("page")) if "page" in args else 1
    # We've switched to using ROWNUM rather than pagination, for performance reasons.
    # This means queries with more than 500 results are invalid.
    #results = results.limit(50).offset((page - 1) * 50).all()
    if current_app.config.get('IS_ORACLE'):
        results = results.filter(
            literal_column("rownum") <= 500
        ).yield_per(50)
    else:
        results = results.limit(500)

    result_fields = [
        "corpNum",
        "corpNme",
        "recognitionDts",
        "corpTypCd",
        "stateTypCd",
        # Due to performance issues, exclude address.
        # 'postalCd'
    ]

    corporations = []
    index = 0
    for row in results:
        if (page-1) * per_page <= index < page * per_page:

            result_dict = {}


            result_dict = {key: getattr(row, convert_to_snake_case(key)) for key in result_fields}
            # Due to performance issues, exclude address.
            result_dict["addr"] = '' #_merge_addr_fields(row)

            corporations.append(result_dict)
        index += 1

    return jsonify({
        "results": corporations,
        "num_results": index
    })


@API.route("/export/")
@jwt.requires_auth
def corporation_search_export():
    """Export a set of Corporation search results to Excel (.xlsx).

    Uses the same parameters as corporation_search().
    """
    account_id = request.headers.get("X-Account-Id", None)
    if not authorized(jwt, account_id):
        return jsonify({"message": "User is not authorized to access Director Search"}), HTTPStatus.UNAUTHORIZED

    # Query string arguments
    args = request.args

    # Fetching results
    results = Corporation.search_corporations(args, include_addr=True)
    if current_app.config.get('IS_ORACLE'):
        results = results.filter(
            literal_column("rownum") <= 500
        ).yield_per(50)
    else:
        results = results.limit(500)

    # Exporting to Excel
    workbook = Workbook()

    export_dir = "/tmp"
    with NamedTemporaryFile(mode="w+b", dir=export_dir, delete=True):

        sheet = workbook.active

        # Sheet headers (first row)
        _ = sheet.cell(column=1, row=1, value="Inc/Reg #")
        _ = sheet.cell(column=2, row=1, value="Entity Type")
        _ = sheet.cell(column=3, row=1, value="Company Name")
        _ = sheet.cell(column=4, row=1, value="Incorporated")
        _ = sheet.cell(column=5, row=1, value="Company Status")
        _ = sheet.cell(column=6, row=1, value="Company Address")
        _ = sheet.cell(column=7, row=1, value="Postal Code")

        index = 2
        for row in results:
            # Corporation.corp_num
            _ = sheet.cell(column=1, row=index, value=row.corp_num)
            # Corporation.corp_typ_cd
            _ = sheet.cell(column=2, row=index, value=row.corp_typ_cd)
            # CorpName.corp_nme
            _ = sheet.cell(column=3, row=index, value=row.corp_nme)
            # Corporation.recognition_dts
            _ = sheet.cell(column=4, row=index, value=row.recognition_dts)
            # CorpOpState.state_typ_cd
            _ = sheet.cell(column=5, row=index, value=row.state_typ_cd)
            # Address.addr_line_1, Address.addr_line_2, Address.addr_line_3
            _ = sheet.cell(column=6, row=index, value=_merge_addr_fields(row))
            # Address.postal_cd
            _ = sheet.cell(column=7, row=index, value=row.postal_cd)
            index += 1

        current_date = datetime.datetime.strftime(datetime.datetime.now(), "%Y-%m-%d %H:%M:%S")
        filename = "Corporation Search Results {date}.xlsx".format(date=current_date)
        full_filename_path = "{dir}/{filename}".format(dir=export_dir, filename=filename)
        workbook.save(filename=full_filename_path)

        return send_from_directory(export_dir, filename, as_attachment=True)


@API.route("/<corp_id>")
@jwt.requires_auth
def corporation(corp_id):
    """Get a single Corporation by corpNum."""
    account_id = request.headers.get("X-Account-Id", None)
    if not authorized(jwt, account_id):
        return jsonify({"message": "User is not authorized to access Director Search"}), HTTPStatus.UNAUTHORIZED

    corp = Corporation.get_corporation_by_id(corp_id)
    if not corp:
        return jsonify({"message": "Corporation with id {} could not be found.".format(corp_id)}), 404

    offices = Office.get_offices_by_corp_id(corp_id)
    names = CorpName.get_corp_name_by_corp_id(corp_id)

    output = {}
    output["corpNum"] = corp.corp_num
    output["transitionDt"] = corp.transition_dt
    output["offices"] = []
    for office in offices:
        output["offices"].append(
            {
                "deliveryAddr": Address.normalize_addr(office.delivery_addr_id),
                "mailingAddr": Address.normalize_addr(office.mailing_addr_id),
                "officeTypCd": _format_office_typ_cd(office.office_typ_cd),
                "emailAddress": office.email_address,
            }
        )

    output["adminEmail"] = corp.admin_email

    output["NAMES"] = []
    for row in names:
        output["NAMES"].append({"name": row.corp_nme})

    return jsonify(output)
