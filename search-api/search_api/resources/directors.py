# Copyright © 2020 Province of British Columbia
#
# Licensed under the Apache License, Version 2.0 (the 'License');
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an 'AS IS' BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""API endpoints for searching for and retrieving information about directors (CorpParties)."""

import datetime
from http import HTTPStatus
import logging
from tempfile import NamedTemporaryFile

from flask import Blueprint, current_app, request, jsonify, send_from_directory
from sqlalchemy.sql import literal_column
from openpyxl import Workbook

from search_api.auth import jwt, authorized
from search_api.models.address import Address
from search_api.models.corp_state import CorpState
from search_api.models.corp_party import CorpParty
from search_api.models.party_type import PartyType
from search_api.models.corp_name import CorpName
from search_api.models.office import Office
from search_api.utils.model_utils import (
    BadSearchValue,
    _get_corp_party_export_column_headers,
    _get_corp_party_export_column_values,
)
from search_api.utils.utils import convert_to_snake_case

logger = logging.getLogger(__name__)
API = Blueprint('DIRECTORS_API', __name__, url_prefix='/api/v1/directors')


@API.route('/')
@jwt.requires_auth
def corpparty_search():
    """Search for CorpParty entities.

    This function takes any number of field triples in the following format:
    - field={field name}
    - operator={'exact', 'contains', 'startswith', 'nicknames', 'similar' or 'endswith'}
    - value={search keyword}

    To include Address or CorpOpState info in the search results, set the additional_cols
    field to 'addr' or 'active', respectively.
    - additional_cols={'none', 'addr', or 'active'}

    In addition, the following arguments are provided:
    - page={page number}
    - sort_type={'asc' or 'dsc'}
    - sort_value={field name to sort results by}
    """
    current_app.logger.info('Starting director search')
    account_id = request.headers.get('X-Account-Id', None)
    if not authorized(jwt, account_id):
        return (
            jsonify({'message': 'User is not authorized to access Director Search'}),
            HTTPStatus.UNAUTHORIZED,
        )

    current_app.logger.info('Authorization check finished; starting query {query}'.format(query=request.url))

    args = request.args
    fields = args.getlist('field')
    additional_cols = args.get('additionalCols')
    try:
        results = CorpParty.search_corp_parties(args)
    except BadSearchValue as e:
        return {'results': [], 'error': 'Invalid search: {}'.format(str(e))}
    current_app.logger.info('Before query')

    # Pagination
    page = int(args.get('page')) if 'page' in args else 1

    per_page = 50
    # Manually paginate results, because flask-sqlalchemy's paginate() method counts the total,
    # which is slow for large tables. This has been addressed in flask-sqlalchemy but is unreleased.
    # Ref: https://github.com/pallets/flask-sqlalchemy/pull/613
    # results = results.limit(per_page).offset((page - 1) * per_page).all()
    # update:
    # We've switched to using ROWNUM rather than pagination, for performance reasons.
    # for benchmarking, dump the query here and copy to benchmark.py
    # This means queries with more than 165 results are invalid.

    if current_app.config.get('IS_ORACLE'):
        results = results.filter(literal_column('rownum') <= 165)  # .yield_per(per_page)
    else:
        results = results.limit(165)

    # for debugging, print out the exact query.
    # from sqlalchemy.dialects import oracle
    # oracle_dialect = oracle.dialect(max_identifier_length=30)
    # raise Exception(results.statement.compile(dialect=oracle_dialect))

    current_app.logger.info('After query')

    result_fields = [
        'corpPartyId',
        'firstNme',
        'middleNme',
        'lastNme',
        'appointmentDt',
        'cessationDt',
        'corpNum',
        'corpNme',
        'partyTypCd',
        'corpAdminEmail',
    ]

    corp_parties = []
    index = 0
    for row in results:
        if (page - 1) * per_page <= index < page * per_page:
            result_dict = {key: getattr(row, convert_to_snake_case(key)) for key in result_fields}
            result_dict['corpPartyId'] = int(result_dict['corpPartyId'])

            additional_result_columns = CorpParty.add_additional_cols_to_search_results(additional_cols, fields, row)

            result_dict = {**result_dict, **additional_result_columns}

            corp_parties.append(result_dict)
        index += 1

    current_app.logger.info('Returning JSON results')

    return jsonify({
        'results': corp_parties,
        'num_results': index
    })


@API.route('/export/')
@jwt.requires_auth
def corpparty_search_export():
    """Export a list of CorpParty search results. Uses the same arguments as corpparty_search()."""
    account_id = request.headers.get('X-Account-Id', None)
    if not authorized(jwt, account_id):
        return (
            jsonify({'message': 'User is not authorized to access Director Search'}),
            HTTPStatus.UNAUTHORIZED,
        )

    # Query string arguments
    args = request.args

    # Fetching results
    results = CorpParty.search_corp_parties(args)
    if current_app.config.get('IS_ORACLE'):
        results = results.filter(literal_column('rownum') <= 165)
    else:
        results = results.limit(165)

    # Exporting to Excel
    workbook = Workbook()

    export_dir = '/tmp'
    with NamedTemporaryFile(mode='w+b', dir=export_dir, delete=True):

        sheet = workbook.active

        for index, column_header in enumerate(_get_corp_party_export_column_headers(args), start=1):
            _ = sheet.cell(column=index, row=1, value=column_header)

        for row_index, row in enumerate(results, start=2):
            for column_index, column_value in enumerate(_get_corp_party_export_column_values(row, args), start=1):
                _ = sheet.cell(column=column_index, row=row_index, value=column_value)

        current_date = datetime.datetime.strftime(datetime.datetime.now(), '%Y-%m-%d %H:%M:%S')
        filename = 'Director Search Results {date}.xlsx'.format(date=current_date)
        workbook.save(filename='{dir}/{filename}'.format(dir=export_dir, filename=filename))

        return send_from_directory(export_dir, filename, as_attachment=True)


@API.route('/<corp_party_id>')
@jwt.requires_auth
def get_corp_party_by_id(corp_party_id):
    """Get a CorpParty by id."""
    account_id = request.headers.get('X-Account-Id', None)
    if not authorized(jwt, account_id):
        return (
            jsonify({'message': 'User is not authorized to access Director Search'}),
            HTTPStatus.UNAUTHORIZED,
        )

    result = CorpParty.get_corporation_info_by_corp_party_id(corp_party_id)

    if not result:
        return jsonify({'message': 'Director with id {} could not be found.'.format(corp_party_id)}), 404

    person = result[0]
    result_dict = {}

    filing_description = CorpParty.get_filing_description_by_corp_party_id(corp_party_id)

    name = CorpName.get_corp_name_by_corp_id(person.corp_num)[0]
    offices = Office.get_offices_by_corp_id(person.corp_num)
    delivery_addr = Address.normalize_addr(person.delivery_addr_id)
    mailing_addr = Address.normalize_addr(person.mailing_addr_id)

    states = CorpState.get_corp_states_by_corp_id(person.corp_num)

    corp_delivery_addr = ';'.join([Address.normalize_addr(office.delivery_addr_id) for office in offices])
    corp_mailing_addr = ';'.join([Address.normalize_addr(office.mailing_addr_id) for office in offices])

    result_dict['corpPartyId'] = int(person.corp_party_id)
    result_dict['firstNme'] = person.first_nme
    result_dict['middleNme'] = person.middle_nme
    result_dict['lastNme'] = person.last_nme
    result_dict['appointmentDt'] = person.appointment_dt
    result_dict['cessationDt'] = person.cessation_dt
    result_dict['corpNum'] = person.corp_num
    result_dict['corpNme'] = name.corp_nme
    result_dict['partyTypCd'] = person.party_typ_cd
    result_dict['partyTypeDesc'] = PartyType.query.filter(
        PartyType.party_typ_cd == person.party_typ_cd).one().short_desc
    result_dict['corpPartyEmail'] = person.email_address
    result_dict['deliveryAddr'] = delivery_addr
    result_dict['mailingAddr'] = mailing_addr
    result_dict['corpDeliveryAddr'] = corp_delivery_addr
    result_dict['corpMailingAddr'] = corp_mailing_addr
    result_dict['corpTypCd'] = result.corp_typ_cd
    result_dict['corpAdminEmail'] = result.corp_admin_email
    result_dict['fullDesc'] = filing_description[0].full_desc if filing_description else None

    result_dict['states'] = [s.as_dict() for s in states]

    offices_held = _get_offices_held_by_corp_party(corp_party_id)

    incorporator_name_types = [
        'APP',
        'ATT',
        'FBO',
        'FCP',
        'FIO',
        'INC',
        'LIQ',
        'PAS',
        'PSA',
        'RAD',
        'RAF',
        'RAO',
        'RCC',
        'RCM',
        'TAA',
        'TAP',
    ]
    if person.party_typ_cd in incorporator_name_types:
        result_dict['businessNme'] = person.business_nme

    return jsonify({**result_dict, **offices_held})


def _get_offices_held_by_corp_party(corp_party_id):
    results = CorpParty.get_offices_held_by_corp_party_id(corp_party_id)

    if len(results) == 0:
        return {'offices': [], 'sameAddr': [], 'sameNameAndCompany': []}

    offices = []
    for row in results:
        result_dict = {}

        result_dict['corpPartyId'] = int(row.corp_party_id)
        result_dict['officerTypCd'] = row.officer_typ_cd
        result_dict['shortDesc'] = row.short_desc
        result_dict['appointmentDt'] = row.appointment_dt
        result_dict['year'] = row.event_timestmp.year if row.event_timestmp else None

        offices.append(result_dict)

    same_addr = CorpParty.get_corp_party_at_same_addr(corp_party_id)
    same_name_and_company = CorpParty.get_corp_party_same_name_at_same_addr(corp_party_id)

    results = {
        'offices': offices,
        'sameAddr': [
            {**s[0].as_dict(), **{'year': int(s[1].year)}}
            for s in same_addr
            if s[0].corp_party_id != int(corp_party_id)
        ],
        'sameNameAndCompany': [
            {**s[0].as_dict(), **{'year': int(s[1].year)}}
            for s in same_name_and_company
            if s[0].corp_party_id != int(corp_party_id)
        ],
    }
    return results
